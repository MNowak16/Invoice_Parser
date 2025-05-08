#!/usr/bin/env python
# coding: utf-8

# In[1]:


# Import libraries 
import fitz
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import os
import re


# In[2]:


# Function to extract the data from a single pdf
def extract_pdf_data(pdf_path):
    doc = fitz.open(pdf_path)

    # List of fields
    fields = ["Date", "Payee", "Fund", "Department", "Account", 
          "INVOICE ACCOUNT OR WINRow2", "INVOICE DATERow2", "DESCRIPTIONRow2", "AMOUNTRow2",
          "INVOICE ACCOUNT OR WINRow3", "INVOICE DATERow3", "DESCRIPTIONRow3", "AMOUNTRow3",
          "INVOICE ACCOUNT OR WINRow4", "INVOICE DATERow4", "DESCRIPTIONRow4", "AMOUNTRow4",
          "INVOICE ACCOUNT OR WINRow5", "INVOICE DATERow5", "DESCRIPTIONRow5", "AMOUNTRow5",
          "INVOICE ACCOUNT OR WINRow6", "INVOICE DATERow6", "DESCRIPTIONRow6", "AMOUNTRow6",
          "INVOICE ACCOUNT OR WINRow7", "INVOICE DATERow7", "DESCRIPTIONRow7", "AMOUNTRow7",
          "INVOICE ACCOUNT OR WINRow8", "INVOICE DATERow8", "DESCRIPTIONRow8", "AMOUNTRow8",
          "INVOICE ACCOUNT OR WINRow9", "INVOICE DATERow9", "DESCRIPTIONRow9", "AMOUNTRow9",
          "INVOICE ACCOUNT OR WINRow10", "INVOICE DATERow10", "DESCRIPTIONRow10", "AMOUNTRow10"]

    # Get the header data
    form_data = {}
    for widget in doc[0].widgets():
        field_name = widget.field_name
        field_value = widget.field_value
        if field_name:
            form_data[field_name] = field_value if field_value else ""

    # Extract header information
    header = {
        "Date": form_data.get("Date"),
        "Payee": form_data.get("Payee"),
        "Fund": form_data.get("Fund"),
        "Department": form_data.get("Department"),
        "Account": form_data.get("Account")
    }

    # Build rows for each invoice using headers as repeating rows
    invoices = []
    last_full_row = None
    
    for i in range(2, 11):  # Rows 2 to 10
        invoice_no = form_data.get(f"INVOICE ACCOUNT OR WINRow{i}", "")
        date = form_data.get(f"INVOICE DATERow{i}", "")
        description = form_data.get(f"DESCRIPTIONRow{i}", "")
        amount = form_data.get(f"AMOUNTRow{i}", "")
        
        if not any([invoice_no, date, amount]) and description and last_full_row:
            # This is a continuation row → add to previous description
            last_full_row["Description"] += " " + description
            continue
    
        if not any([invoice_no, date, description, amount]):
            continue  # Skip completely blank rows
    
        # Build new full row with header and invoice data
        row = {
            **header,  # Include the header values in each row
            "Invoice_No": invoice_no,
            "Invoice_Date": date,
            "Description": description,
            "Amount": amount
        }
        
        invoices.append(row)
        last_full_row = row  # Keep reference for possible continuation

    return invoices


# In[ ]:


# Function to parse and format date manually
def parse_and_format_date(date_str):
    try:
        if not isinstance(date_str, str):
            return pd.NaT

        date_str = date_str.strip()

        # Split using regex (either '/' or '.')
        parts = re.split(r'[/.]', date_str)
        if len(parts) != 3:
            return pd.NaT

        month, day, year = parts

        # Zero pad month and day
        month = month.zfill(2)
        day = day.zfill(2)

        # Convert 2-digit years
        if len(year) == 2:
            year = '20' + year

        cleaned_date = f"{month}/{day}/{year}"

        # Attempt final parse to datetime
        return pd.to_datetime(cleaned_date, format='%m/%d/%Y', errors='coerce')

    except Exception:
        return pd.NaT


# In[4]:


# Function to process all PDFs in a folder
def process_pdfs_in_folder():
    # Open the folder selection dialog
    root = tk.Tk()
    root.withdraw()  # Hide the Tkinter root window
    folder_path = filedialog.askdirectory(title="Select Folder Containing PDFs")

    if not folder_path:
        print("No folder selected. Exiting...")
        return

    # Process all PDFs in the selected folder
    all_invoices = []
    for file_name in os.listdir(folder_path):
        if file_name.lower().endswith(".pdf"):
            pdf_path = os.path.join(folder_path, file_name)
            print(f"Processing {pdf_path}...")
            invoices = extract_pdf_data(pdf_path)
            all_invoices.extend(invoices)

    # Convert to DataFrame
    df = pd.DataFrame(all_invoices)

    # Apply the date parsing function to the Invoice_Date column
    df['Invoice_Date'] = df['Invoice_Date'].apply(parse_and_format_date)
    
    # Convert data types
    df["Date"] = pd.to_datetime(df["Date"], format='%m/%d/%Y')
    df["Invoice_Date"] = pd.to_datetime(df["Invoice_Date"], format='%m/%d/%Y')
    #df["Amount"] = df["Amount"].astype(float)

    # Excel file path
    excel_path = os.path.join(folder_path, "Invoice_Output.xlsx")

    # Check if file exists
    if os.path.exists(excel_path):
        # Load the workbook to get the last row
        book = load_workbook(excel_path)
        sheet = book.active
        start_row = sheet.max_row  # Determine where to start writing

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=start_row)
    else:
        # First-time write: include headers
        df.to_excel(excel_path, index=False)

    print(f"Data extraction complete. Saved to '{excel_path}'.")


# In[ ]:


if __name__ == "__main__":
    process_pdfs_in_folder()

